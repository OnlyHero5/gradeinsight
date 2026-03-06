from __future__ import annotations

import re
import warnings
from dataclasses import dataclass
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .exam_identity import build_exam_identity, looks_like_exam_text
from .number_utils import to_decimal, to_int

QUESTION_KEY_PATTERN = re.compile(r"^\d+(?:\(\d+\))?$")
TOTAL_SCORE_HEADERS = ("总分", "得分")


@dataclass(frozen=True)
class ParsedStudentRow:
    name: str
    external_id: str
    admission_ticket: str
    custom_exam_id: str
    class_name: str
    total_score: Decimal | None
    total_raw: str
    objective_score: Decimal | None
    subjective_score: Decimal | None
    rank_in_class: int | None
    excluded_from_stats: bool
    exclude_reason: str
    question_scores: dict[str, Decimal | None]


@dataclass(frozen=True)
class ParsedExam:
    question_keys: list[str]
    rows: list[ParsedStudentRow]
    student_count: int
    excluded_from_stats_count: int
    mismatched_id_count: int
    identity_key: str
    identity_label: str


def parse_exam_xlsx(file_bytes: bytes, source_filename: str = "") -> ParsedExam:
    workbook = _load_workbook(file_bytes)
    try:
        return _parse_detail_workbook(workbook, source_filename=source_filename)
    except ValueError:
        try:
            return _parse_total_only_workbook(workbook, source_filename=source_filename)
        except ValueError as exc:
            raise ValueError("无法识别 .xlsx 结构：未找到题目明细结构，也未找到可用的汇总/简表结构") from exc


def _load_workbook(file_bytes: bytes) -> Workbook:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        return load_workbook(BytesIO(file_bytes), data_only=True)


def _parse_detail_workbook(workbook: Workbook, source_filename: str) -> ParsedExam:
    for worksheet in _iter_detail_candidate_sheets(workbook):
        try:
            header_row = _find_detail_header_row(worksheet)
            header_map = _build_header_map(worksheet, header_row)
            question_columns = _extract_question_columns(header_map)
            rows, excluded_count, mismatch_count = _parse_detail_rows(
                worksheet,
                header_row,
                header_map,
                question_columns,
            )
            identity = build_exam_identity(
                title_candidates=_collect_workbook_title_candidates(workbook),
                source_filename=source_filename,
                fallback_class_name=rows[0].class_name if rows else "",
            )
            return ParsedExam(
                question_keys=[question_key for question_key, _ in question_columns],
                rows=rows,
                student_count=len(rows),
                excluded_from_stats_count=excluded_count,
                mismatched_id_count=mismatch_count,
                identity_key=identity.key,
                identity_label=identity.label,
            )
        except ValueError:
            continue

    raise ValueError("未找到题目明细结构")


def _parse_total_only_workbook(workbook: Workbook, source_filename: str) -> ParsedExam:
    for worksheet in _iter_total_only_candidate_sheets(workbook):
        try:
            header_row = _find_total_only_header_row(worksheet)
            header_map = _build_header_map(worksheet, header_row)
            rows, excluded_count, mismatch_count = _parse_total_only_rows(worksheet, header_row, header_map)
            identity = build_exam_identity(
                title_candidates=_collect_workbook_title_candidates(workbook),
                source_filename=source_filename,
                fallback_class_name=rows[0].class_name if rows else "",
            )
            return ParsedExam(
                question_keys=[],
                rows=rows,
                student_count=len(rows),
                excluded_from_stats_count=excluded_count,
                mismatched_id_count=mismatch_count,
                identity_key=identity.key,
                identity_label=identity.label,
            )
        except ValueError:
            continue

    raise ValueError("未找到可用的汇总/简表结构")


def _iter_detail_candidate_sheets(workbook: Workbook) -> list[Worksheet]:
    sheets: list[Worksheet] = []
    if "得分明细" in workbook.sheetnames:
        sheets.append(workbook["得分明细"])
    sheets.extend([worksheet for worksheet in workbook.worksheets if worksheet not in sheets])
    return sheets


def _iter_total_only_candidate_sheets(workbook: Workbook) -> list[Worksheet]:
    sheets: list[Worksheet] = []
    preferred_titles = ("班级英语成绩汇总", "班级成绩汇总", "班级英语成绩简表", "班级成绩简表")

    for title in preferred_titles:
        if title in workbook.sheetnames:
            sheets.append(workbook[title])

    seen_titles = {sheet.title for sheet in sheets}
    for worksheet in workbook.worksheets:
        if worksheet.title not in seen_titles and "汇总" in worksheet.title:
            sheets.append(worksheet)
            seen_titles.add(worksheet.title)
    for worksheet in workbook.worksheets:
        if worksheet.title not in seen_titles and "简表" in worksheet.title:
            sheets.append(worksheet)
            seen_titles.add(worksheet.title)
    for worksheet in workbook.worksheets:
        if worksheet.title not in seen_titles:
            sheets.append(worksheet)
            seen_titles.add(worksheet.title)

    return sheets


def _collect_workbook_title_candidates(workbook: Workbook) -> list[str]:
    candidates: list[str] = []
    for worksheet in workbook.worksheets:
        for row_index in range(1, min(worksheet.max_row, 3) + 1):
            values = [
                str(worksheet.cell(row=row_index, column=column).value or "").strip()
                for column in range(1, min(worksheet.max_column, 12) + 1)
            ]
            joined = " ".join(item for item in values if item)
            if joined and looks_like_exam_text(joined):
                candidates.append(joined)
    return candidates


def _find_detail_header_row(worksheet: Worksheet) -> int:
    max_scan_rows = min(worksheet.max_row, 12)
    for row in range(1, max_scan_rows + 1):
        values = {str(worksheet.cell(row=row, column=col).value or "").strip() for col in range(1, 80)}
        if {"姓名", "总分"}.issubset(values):
            return row

    raise ValueError("无法定位明细考试表头")


def _find_total_only_header_row(worksheet: Worksheet) -> int:
    max_scan_rows = min(worksheet.max_row, 12)
    for row in range(1, max_scan_rows + 1):
        values = {str(worksheet.cell(row=row, column=col).value or "").strip() for col in range(1, 80)}
        if "姓名" in values and any(header in values for header in TOTAL_SCORE_HEADERS):
            return row

    raise ValueError("无法定位仅总分考试表头")


def _build_header_map(worksheet: Worksheet, header_row: int) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col in range(1, worksheet.max_column + 1):
        header = str(worksheet.cell(row=header_row, column=col).value or "").strip()
        if header:
            mapping[header] = col
    return mapping


def _extract_question_columns(header_map: dict[str, int]) -> list[tuple[str, int]]:
    question_items = [
        (header, col)
        for header, col in sorted(header_map.items(), key=lambda item: item[1])
        if QUESTION_KEY_PATTERN.match(header)
    ]
    if not question_items:
        raise ValueError("未找到题号列（如 1、51、51(3)）")
    return question_items


def _find_total_column(header_map: dict[str, int]) -> int | None:
    for header in TOTAL_SCORE_HEADERS:
        if header in header_map:
            return header_map[header]
    return None


def _parse_detail_rows(
    worksheet: Worksheet,
    header_row: int,
    header_map: dict[str, int],
    question_columns: list[tuple[str, int]],
) -> tuple[list[ParsedStudentRow], int, int]:
    rows: list[ParsedStudentRow] = []
    excluded_count = 0
    mismatch_count = 0

    for row_index in range(header_row + 1, worksheet.max_row + 1):
        name = _read_text(worksheet, row_index, header_map.get("姓名"))
        if not name:
            continue

        admission_ticket = _read_text(worksheet, row_index, header_map.get("准考证号"))
        custom_exam_id = _read_text(worksheet, row_index, header_map.get("自定义考号"))
        external_id = admission_ticket or custom_exam_id
        if not external_id:
            continue

        if admission_ticket and custom_exam_id and admission_ticket != custom_exam_id:
            mismatch_count += 1

        total_raw_value = worksheet.cell(row=row_index, column=header_map.get("总分", 0)).value
        total_raw = "" if total_raw_value is None else str(total_raw_value).strip()
        total_score = to_decimal(total_raw_value)
        excluded = total_score is None
        exclude_reason = "non_numeric_total_score" if excluded and total_raw else ""

        if excluded:
            excluded_count += 1

        question_scores = {
            question_key: to_decimal(worksheet.cell(row=row_index, column=column).value)
            for question_key, column in question_columns
        }

        rows.append(
            ParsedStudentRow(
                name=name,
                external_id=external_id,
                admission_ticket=admission_ticket,
                custom_exam_id=custom_exam_id,
                class_name=_read_text(worksheet, row_index, header_map.get("班级")),
                total_score=total_score,
                total_raw=total_raw,
                objective_score=to_decimal(_read_value(worksheet, row_index, header_map.get("客观分"))),
                subjective_score=to_decimal(_read_value(worksheet, row_index, header_map.get("主观分"))),
                rank_in_class=to_int(_read_value(worksheet, row_index, header_map.get("班次"))),
                excluded_from_stats=excluded,
                exclude_reason=exclude_reason,
                question_scores=question_scores,
            )
        )

    return rows, excluded_count, mismatch_count


def _parse_total_only_rows(
    worksheet: Worksheet,
    header_row: int,
    header_map: dict[str, int],
) -> tuple[list[ParsedStudentRow], int, int]:
    total_column = _find_total_column(header_map)
    if total_column is None or "姓名" not in header_map:
        raise ValueError("缺少姓名或总分/得分列")

    rows: list[ParsedStudentRow] = []
    excluded_count = 0
    mismatch_count = 0

    for row_index in range(header_row + 1, worksheet.max_row + 1):
        name = _read_text(worksheet, row_index, header_map.get("姓名"))
        if not name:
            continue

        admission_ticket = _read_text(worksheet, row_index, header_map.get("准考证号"))
        custom_exam_id = _read_text(worksheet, row_index, header_map.get("自定义考号"))
        external_id = admission_ticket or custom_exam_id

        if admission_ticket and custom_exam_id and admission_ticket != custom_exam_id:
            mismatch_count += 1

        total_raw_value = _read_value(worksheet, row_index, total_column)
        total_raw = "" if total_raw_value is None else str(total_raw_value).strip()
        total_score = to_decimal(total_raw_value)
        excluded = total_score is None
        exclude_reason = "non_numeric_total_score" if excluded and total_raw else ""

        if excluded:
            excluded_count += 1

        rows.append(
            ParsedStudentRow(
                name=name,
                external_id=external_id,
                admission_ticket=admission_ticket,
                custom_exam_id=custom_exam_id,
                class_name=_read_text(worksheet, row_index, header_map.get("班级")),
                total_score=total_score,
                total_raw=total_raw,
                objective_score=None,
                subjective_score=None,
                rank_in_class=to_int(_read_value(worksheet, row_index, header_map.get("班次"))),
                excluded_from_stats=excluded,
                exclude_reason=exclude_reason,
                question_scores={},
            )
        )

    if not rows:
        raise ValueError("仅总分考试未解析到学生数据")

    return rows, excluded_count, mismatch_count


def _read_text(worksheet: Worksheet, row: int, column: int | None) -> str:
    value = _read_value(worksheet, row, column)
    return "" if value is None else str(value).strip()


def _read_value(worksheet: Worksheet, row: int, column: int | None) -> Any:
    if column is None:
        return None
    return worksheet.cell(row=row, column=column).value
